"""Build a review-friendly xlsx of all kit protocol sequences (Old vs New)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Protocol Sequences"

header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="1F4E78")
group_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
group_fill = PatternFill("solid", start_color="2E75B6")
body_font = Font(name="Arial", size=10)
new_col_fill = PatternFill("solid", start_color="FFF2CC")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    "DiagnosisKey",
    "Condition",
    "Old Sequence (current backend)",
    "New Sequence (your edits)",
    "Rationale",
    "Runtime overlays / notes",
]
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[1].height = 28

data = [
    ("TELOGEN EFFLUVIUM", "TE_STRESS", "TE — Stress / Anxiety / Depression",
     "HAIR FACT TE GOLD → PHENOTYPE INFLAMATION",
     "Active shedding arrested first; inflammation cleared in Ph2",
     "TE GOLD stripped if duration > 3m / thinning-only / regrow / breastfeeding / GI GOLD present"),
    ("TELOGEN EFFLUVIUM", "TE_NUTRITION", "TE — Nutritional Deficiency",
     "HAIR FACT TE GOLD → PHENOTYPE INFLAMATION",
     "TE arrest first; residual nutritional-stress inflammation cleared",
     "PRO IMMUNE injected when needsImmune signal fires"),
    ("TELOGEN EFFLUVIUM", "TE_POSTPREG", "TE — Post-partum / Breastfeeding",
     "LACTIHEALTH → HAIR FACT TE GOLD → PRO IMMUNE GOLD",
     "Lactation demand first; TE in Ph2; immunity rebuild",
     "TE GOLD STRIPPED if still feeding (LACTIHEALTH supersedes)"),
    ("TELOGEN EFFLUVIUM", "TE_DELIVERY", "TE — Post-delivery, Not Feeding",
     "HAIR FACT TE GOLD → PRO IMMUNE GOLD",
     "Post-natal TE first; immunity rebuild", ""),
    ("TELOGEN EFFLUVIUM", "TE_ILLNESS", "TE — Illness / Surgery / Medication",
     "PRO IMMUNE GOLD → HAIR FACT TE GOLD → PHENOTYPE INFLAMATION",
     "Immunity cleared first; then TE; then inflammation", ""),

    ("PATTERN HAIR LOSS", "AGA_MALE_123", "MPHL — Grade 1-2-3",
     "HAIR FACT TE GOLD → PHENOTYPE → MPHL → PRO FACT META B → PRO IMMUNE GOLD",
     "5-phase: arrest → clear → correct DHT → metabolic → regrow",
     "Pattern kit upgrades to MPHL PLUS when isGrade45"),
    ("PATTERN HAIR LOSS", "AGA_MALE_45", "MPHL — Grade 4-5 Advanced",
     "PRO IMMUNE GOLD → MPHL → PHENOTYPE INFLAMATION",
     "Rescue environment → DHT block → clear microenvironment",
     "Resolves to MPHL PLUS; bald areas not reversible (prognosis addendum)"),
    ("PATTERN HAIR LOSS", "AGA_FEMALE_123", "FPHL — Grade 1-2-3",
     "HAIR FACT TE GOLD → PHENOTYPE → FPHL → PRO FACT META B → PRO IMMUNE GOLD",
     "5-phase female pattern", ""),
    ("PATTERN HAIR LOSS", "AGA_FEMALE_45", "FPHL — Grade 4-5 Advanced",
     "PRO IMMUNE GOLD → FPHL → PHENOTYPE INFLAMATION",
     "3-phase advanced female pattern",
     "Resolves to FPHL PLUS"),

    ("PCOS", "PCOS_ONLY", "PCOS — Hormonal Hair Loss",
     "PRO FACT META B PCOS → PHENOTYPE INFLAMATION → HAIR FACT TE GOLD",
     "PCOS-specific kit covers androgen + insulin in one; F-PCOS-1 retired",
     "Upgrades when diabetes/pre-diabetes signal present"),
    ("PCOS", "PCOS_OBESITY", "PCOS + Metabolic (Insulin Resistance)",
     "PRO FACT META B PCOS → PHENOTYPE INFLAMATION",
     "Single primary kit covers AMPK+androgen; no F-PCOS-1", ""),

    ("THYROID", "THYROID_HYPO", "Hypothyroidism",
     "PRO FACT META B HYPOTHYROID → PHENOTYPE INFLAMATION → HAIR FACT TE GOLD",
     "Thyroid first; inflammation; concurrent TE",
     "OVERRIDE: obesity/sedentary present → swap to PRO FACT META B (2-condition kit)"),
    ("THYROID", "THYROID_HYPER", "Hyperthyroidism",
     "PRO FACT THYROID CARE → HAIR FACT TE GOLD → PRO IMMUNE GOLD",
     "Metabolic overactivity → TE → immune rebuild", ""),

    ("ALOPECIA AREATA", "ALOPECIA_AREATA", "Alopecia Areata (Autoimmune)",
     "HAIR FACT ALOPECIA AREATA → PRO FACT META B → PHENOTYPE → PRO IMMUNE GOLD",
     "4-phase autoimmune protocol",
     "Grade question skipped when AA present"),

    ("MENOPAUSE CONTINUUM", "PERI_MENOPAUSE", "Peri-Menopause",
     "HAIR FACT PERI MENOPAUSE → FPHL → HAIR FACT TE GOLD → PRO IMMUNE GOLD",
     "PERI kit stabilises; FPHL; TE arrest; immune last",
     "PRO FACT META B injected when metabolic signals; TE GOLD strip when PERI present"),
    ("MENOPAUSE CONTINUUM", "MENOPAUSE", "Menopause",
     "PRO FACT META B MENOPAUSE → FPHL → HAIR FACT TE GOLD",
     "META B MENO corrects oestrogen decline; FPHL; TE",
     "?? Should this also supersede TE GOLD"),
    ("MENOPAUSE CONTINUUM", "POST_MENOPAUSE", "Post-Menopause",
     "PRO FACT META B POSTMENOPAUSE → FPHL → HAIR FACT TE GOLD → PRO IMMUNE GOLD",
     "Post-meno metabolic shift; FPHL; TE; immune",
     "?? Should also supersede TE GOLD; HAIR FACT HYPOTHYROID POST M - 3 variant?"),

    ("ENDOMETRIOSIS", "ENDOMETRIOSIS", "Endometriosis",
     "FH WELL 3 → PHENOTYPE INFLAMATION → PRO IMMUNE GOLD",
     "Endometriotic balance → inflammation → immune", ""),

    ("PREGNANCY", "PREGNANCY", "Pregnancy",
     "HEALTHY - 9",
     "LOCKED single kit — no other kits during pregnancy",
     "Absolute lock"),

    ("IRON DEFICIENCY", "IRON_DEFICIENCY", "Iron Deficiency / Anaemia",
     "IRON UP GOLD → HAIR FACT TE GOLD → PRO IMMUNE GOLD",
     "Ferritin first (non-negotiable); TE; immune",
     "IRON UP injected for heavy bleeding 18-50 females"),

    ("SCALP INFLAMMATION", "SCALP_INFLAM", "Scalp Inflammation (Dandruff/Oily/Seborrhoea)",
     "PHENOTYPE INFLAMATION → AGA_GENDER",
     "Inflammation cleared before pattern correction",
     "AGA_GENDER resolves to MPHL/FPHL at runtime"),

    ("HAIR BREAKAGE", "HAIR_BREAKAGE", "Hair Breakage (Chemical/Heat/Hard water)",
     "HAIR FACT HAIR BREAKAGE REPAIR(HBR) → PHENOTYPE INFLAMATION",
     "Shaft repair → scalp inflammation",
     "HBR only when broken/short hairs confirmed; heat/chemical alone insufficient"),

    ("LIFESTYLE", "NIGHT_SHIFT", "Night Shift / Circadian Disruption",
     "HAIR FACT NIGHT SHIFT → HAIR FACT TE GOLD → PHENOTYPE INFLAMATION",
     "Circadian → TE → inflammation", ""),
    ("LIFESTYLE", "FREQUENT_FLYING", "Frequent Flying",
     "HAIR FACT FREQUENT FLYERS → HAIR FACT TE GOLD → PHENOTYPE INFLAMATION",
     "Travel stress → TE → inflammation", ""),
    ("LIFESTYLE", "OXIDATIVE", "Smoking / Alcohol / Vaping / Oxidative Stress",
     "PHENOTYPE INFLAMATION → OXIDATIVE STRESS → PRO IMMUNE GOLD",
     "PHENOTYPE enables OXIDATIVE kit; immune rebuild",
     "Standalone OXIDATIVE STRESS injected when oxidativeCount >= 2"),
    ("LIFESTYLE", "WEIGHT_LOSS", "Sudden Weight Loss (Crash diet / GLP-1 / Illness)",
     "RAPID WEIGHT LOSS SHIELD → HAIR FACT TE GOLD → PRO IMMUNE GOLD",
     "Shield first; TE reduced scope; immune",
     "GLP-1 within 6m → Shield forced Pos 0; GLP-1 after 6m → Shield Pos 1"),

    ("BEHAVIOURAL", "TTM", "Trichotillomania (OCD)",
     "HAIR FACT TTM (OCD) → HAIR FACT TE GOLD → PHENOTYPE INFLAMATION",
     "OCD-driven pulling addressed; TE; inflammation", ""),

    ("CHRONIC / METABOLIC", "CHRONIC_MEDICAL", "Chronic Medical Condition (BP / other)",
     "PRO IMMUNE GOLD → PHENOTYPE INFLAMATION",
     "Immune resilience; chronic inflammation",
     "META B added only when true metabolic signal"),
    ("CHRONIC / METABOLIC", "DIABETES", "Diabetes / Pre-Diabetes",
     "PRO FACT META B → PHENOTYPE INFLAMATION → PRO IMMUNE GOLD",
     "Insulin/AMPK correction; inflammation; immune",
     "Upgrades PCOS_ONLY → PRO FACT META B PCOS"),

    ("GUT", "GUT_ISSUES", "Gut Issues (GERD/IBS/Bloating/Crohn)",
     "PRO FACT GI GOLD → PHENOTYPE → PRO IMMUNE GOLD → HAIR FACT TE GOLD",
     "Gut-hair axis first; inflammation; immune; TE last",
     "GI GOLD supersedes TE GOLD — strips TE when both present"),
    ("GUT", "MOUTH_ULCERS", "Mouth / Tongue Ulcers (Immune-Gut)",
     "PHENOTYPE INFLAMATION → PRO IMMUNE GOLD → OXIDATIVE STRESS",
     "Inflammation; immune; oxidative",
     "Score only when gut signal also fires"),

    ("MULTI-FACTORIAL", "MULTI", "Multi-factorial (Stress + Pattern)",
     "PHENOTYPE INFLAMATION → HAIR FACT TE GOLD → AGA_GENDER",
     "Inflammation always Ph1; TE; pattern last",
     "GI GOLD removed from base — injected only if GERD/IBS/Acid/Crohn confirmed"),

    ("EARLY GREYING", "EARLY_GREY", "Early / Premature Greying",
     "EARLY GREYING CARE GOLD → OXIDATIVE STRESS → PRO IMMUNE GOLD",
     "Root cause; free radicals; immune protection",
     "Sole concern → absolute lock; co-condition → kit appended at end"),

    ("REGROW ONLY", "REGROW_ONLY", "Hair Regrowth — Shedding Stopped",
     "PRO IMMUNE GOLD → PHENOTYPE INFLAMATION → PRO FACT META B",
     "Anagen re-entry; inflammation; metabolic terrain",
     "TE GOLD NEVER prescribed; AGA branch keeps PRO FACT META B + PRO IMMUNE + PHENOTYPE if present"),
]

row = 2
current_group = None
for group, key, label, old, rationale, notes in data:
    if group != current_group:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        gc = ws.cell(row=row, column=1, value=group)
        gc.font = group_font
        gc.fill = group_fill
        gc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        current_group = group
    vals = [key, label, old, "", rationale, notes]
    for col_idx, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=col_idx, value=v)
        c.font = body_font
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = border
        if col_idx == 4:
            c.fill = new_col_fill
        if col_idx == 1:
            c.font = Font(name="Arial", size=10, bold=True)
    ws.row_dimensions[row].height = 42
    row += 1

widths = {"A": 22, "B": 36, "C": 56, "D": 56, "E": 50, "F": 46}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# ── Tab 2: Runtime rules ──────────────────────────────────────────────────────
rules_ws = wb.create_sheet("Runtime rules (read-only)")
rules_headers = ["Rule", "When it fires", "Effect on protocol"]
for col_idx, h in enumerate(rules_headers, 1):
    c = rules_ws.cell(row=1, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border
rules_ws.row_dimensions[1].height = 28

rules_data = [
    ("PHENOTYPE INFLAMATION injection",
     "Scalp signals: Dandruff/Redness/Boils/Burning/Flaking/Oily+Itching/Allergies/Asthma/Skin rash/Alopecia Areata/Smoking/Alcohol",
     "Adds PHENOTYPE INFLAMATION when not already in protocol"),
    ("PRO IMMUNE injection",
     "Frequent infections/Allergies/Asthma/Skin rash/Mouth ulcers/Genetics-after-30/Nutritional/Medication/Illness/Surgery",
     "Adds PRO IMMUNE GOLD (or VEG) — anchored to LAST position by proImmuneLastRule"),
    ("OXIDATIVE STRESS injection",
     "oxidativeCount >= 2 (combo of Smoking + Alcohol + Asthma)",
     "Adds OXIDATIVE STRESS as standalone phase when PHENOTYPE not handling it"),
    ("IRON UP injection",
     "Heavy bleeding periods (female 18-50)",
     "Adds IRON UP GOLD to Phase 1"),
    ("HBR injection",
     "Hard water OR (Heat/Chemical + confirmed broken/short hairs)",
     "Adds HAIR FACT HAIR BREAKAGE REPAIR(HBR)"),
    ("LACTIHEALTH injection",
     "Post-partum + still feeding",
     "Adds LACTIHEALTH to Phase 1; strips TE GOLD"),
    ("GI GOLD injection",
     "GERD / IBS / Acid reflux / Crohn at Phase 1 only (NOT Bloating/Constipation/Indigestion)",
     "Adds PRO FACT GI GOLD; supersedes TE GOLD when both present"),
    ("Menopause continuum injection",
     "Age 40-54 + peri OR age 45-54 + meno OR age ≥ 55 + post",
     "Adds PERI/MENO/POST kit; strips conflicting kits (e.g. META B PCOS at age ≥ 55)"),
    ("GLP-1 precedence",
     "GLP-1 within 6 months OR after 6 months in cause",
     "Forces RAPID WEIGHT LOSS SHIELD to Position 0 (early) or Position 1 (late)"),
    ("TE GOLD acute window",
     "Duration 1–3 months + visible fall + not regrow + not breastfeeding",
     "Promotes HAIR FACT TE GOLD to Phase 1"),
    ("TE GOLD strip — duration > 3m",
     "Duration > 3 months",
     "Removes TE GOLD entirely"),
    ("TE GOLD strip — thinning only",
     'count = "Just thinning, no visible fall"',
     "Removes TE GOLD entirely (v39 rule)"),
    ("TE GOLD strip — regrow goal",
     "goal = regrow-only (no active shedding)",
     "Removes TE GOLD; ensures PRO IMMUNE"),
    ("TE GOLD strip — breastfeeding",
     "Post-partum + still feeding",
     "Removes TE GOLD (LACTIHEALTH supersedes)"),
    ("TE GOLD strip — GI GOLD",
     "GI GOLD present",
     "Removes TE GOLD (gut-axis covers nutrient pathway)"),
    ("TE GOLD strip — PERI MENOPAUSE",
     "HAIR FACT PERI MENOPAUSE present",
     "Removes TE GOLD"),
    ("AGA absolute lock (Grade 4/5)",
     "isGrade45 + age ≥ 20",
     "Overrides all scoring → AGA_MALE_45 or AGA_FEMALE_45; pattern kit becomes MPHL PLUS / FPHL PLUS"),
    ("Early grey lock",
     "Sole concern = Early greying",
     "Locks to EARLY_GREY protocol"),
    ("Pregnancy lock",
     "isPregnant",
     "Locks to HEALTHY-9 single kit"),
    ("PRO IMMUNE last",
     "PRO IMMUNE in protocol",
     "Re-anchors PRO IMMUNE to final phase"),
    ("Active shedding promotion",
     "hasActiveShedding flag (1-3m or 3-6m OR 50-100/100+ count OR stress/nutrition causes)",
     "Promotes TE GOLD to Phase 1"),
    ("Kit cap calculation",
     "Smoking/Vaping + active signal count",
     "Adjusts max kit count (typically 5-7)"),
]
for i, (rule, when, effect) in enumerate(rules_data, start=2):
    for col_idx, v in enumerate([rule, when, effect], 1):
        c = rules_ws.cell(row=i, column=col_idx, value=v)
        c.font = body_font
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = border
        if col_idx == 1:
            c.font = Font(name="Arial", size=10, bold=True)
    rules_ws.row_dimensions[i].height = 42
rules_ws.column_dimensions["A"].width = 36
rules_ws.column_dimensions["B"].width = 58
rules_ws.column_dimensions["C"].width = 60
rules_ws.freeze_panes = "A2"

# ── Tab 3: How to use ─────────────────────────────────────────────────────────
how_ws = wb.create_sheet("How to use")
how_ws["A1"] = "How to use this workbook"
how_ws["A1"].font = Font(name="Arial", size=14, bold=True)
notes = [
    "",
    "Tab 1 — Protocol Sequences",
    "  • Old Sequence column = current backend behaviour (READ-ONLY for your reference)",
    "  • New Sequence column (yellow) = edit here. Use → to separate phases.",
    "  • Leave New Sequence blank for any row that should NOT change.",
    "  • Add new rows below the last row of any group if you want to introduce a new diagnosis.",
    "",
    "Tab 2 — Runtime rules",
    "  • These rules run AFTER the base sequence. They inject, strip, or re-order kits.",
    "  • If you want a rule changed, note it in the New Sequence column with a '+ rule:' tag.",
    "    Example: '+ rule: strip TE GOLD when POST_MENOPAUSE kit present'",
    "",
    "When you send the file back:",
    "  1. I will diff Old vs New per row",
    "  2. Apply each non-blank New Sequence to protocolSequencer.ts",
    "  3. Add/modify any rules you flagged on the Runtime tab",
    "  4. Re-run the replay corpus to surface any cases that flip",
    "",
    "Conventions:",
    "  • Use canonical kit names exactly as they appear in the Old column.",
    "  • VEG variants are auto-resolved at runtime — do NOT enumerate them.",
    "  • AGA_GENDER is a placeholder that resolves to MPHL or FPHL based on sex.",
]
for i, t in enumerate(notes, start=2):
    c = how_ws.cell(row=i, column=1, value=t)
    if t.startswith("Tab"):
        c.font = Font(name="Arial", size=11, bold=True, color="1F4E78")
    else:
        c.font = Font(name="Arial", size=10, bold=(t != "" and not t.startswith(" ")))
how_ws.column_dimensions["A"].width = 100

os.makedirs("outputs", exist_ok=True)
wb.save("outputs/protocol_sequences_review.xlsx")
print("saved outputs/protocol_sequences_review.xlsx")
